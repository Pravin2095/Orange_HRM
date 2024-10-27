import openpyxl


def getRowCount(file, sheetname):
    Book = openpyxl.load_workbook(file)
    Sheet = Book[sheetname]
    return Sheet.max_row


def readData(file, sheetname, rownumber, columnnumber):
    Book = openpyxl.load_workbook(file)
    Sheet = Book[sheetname]
    return Sheet.cell(row=rownumber, column=columnnumber).value


def writeData(file, sheetname, rownumber, columnnumber, data):
    Book = openpyxl.load_workbook(file)
    Sheet = Book[sheetname]
    Sheet.cell(row=rownumber, column=columnnumber).value = data
    Book.save(file)
